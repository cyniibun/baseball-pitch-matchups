import pandas as pd
import streamlit as st
import numpy as np
import io

st.set_page_config(page_title="Pitcher vs Batter Analyzer", layout="wide")
st.title("⚾ Pitcher vs Batter Matchup Analyzer")

st.markdown("""
Enter the pitch-specific statistics for both the **pitcher** and the **batter** below. After submitting, the app will generate a detailed comparison table including strikeout and contact metrics, and highlight deltas:

- 🔴 Red = Advantage to Pitcher  
- 🟢 Green = Advantage to Batter
""")

# Input forms
with st.expander("Pitcher Stats Input"):
    pitcher_df = st.experimental_data_editor(
        pd.DataFrame({
            "Pitch": [],
            "K%": [],
            "Whiff%": [],
            "PutAway%": [],
            "OBA": [],
            "BA": [],
            "SLG": []
        }),
        num_rows="dynamic",
        key="pitcher_stats"
    )

with st.expander("Batter Stats Input"):
    batter_df = st.experimental_data_editor(
        pd.DataFrame({
            "Pitch": [],
            "K%": [],
            "Whiff%": [],
            "PutAway%": [],
            "OBA": [],
            "BA": [],
            "SLG": []
        }),
        num_rows="dynamic",
        key="batter_stats"
    )

if st.button("Generate Matchup Table"):
    try:
        # Merge and calculate deltas
        merged = pd.merge(pitcher_df, batter_df, on="Pitch", suffixes=("_Pitcher", "_Batter"))

        for stat in ["K%", "Whiff%", "PutAway%", "OBA", "BA", "SLG"]:
            merged[f"{stat} Delta"] = merged[f"{stat}_Pitcher"] - merged[f"{stat}_Batter"]

        # Conditional formatting using Styler
        def color_deltas(val):
            if pd.isnull(val): return ""
            if val <= -45: return "background-color:#990000; color:white"  # Dark red
            elif val <= -20: return "background-color:#e06666"  # Medium red
            elif val < 0: return "background-color:#f4cccc"  # Light red
            elif val <= 20: return "background-color:#d9ead3"  # Light green
            elif val <= 45: return "background-color:#93c47d"  # Medium green
            else: return "background-color:#38761d; color:white"  # Dark green

        stat_deltas = [col for col in merged.columns if "Delta" in col]
        styled = merged.style.applymap(color_deltas, subset=stat_deltas)

        st.subheader("📊 Matchup Analysis Table")
        st.dataframe(styled, use_container_width=True)

        # Export options
        st.download_button(
            label="📥 Download Table as CSV",
            data=merged.to_csv(index=False).encode("utf-8"),
            file_name="matchup_analysis.csv",
            mime="text/csv"
        )

        # Excel download with formatting (optional)
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.table import Table, TableStyleInfo

        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matchup"

        for r in dataframe_to_rows(merged, index=False, header=True):
            ws.append(r)

        # Apply simple header styling
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        wb.save(output)
        st.download_button(
            label="📊 Download as Excel (.xlsx)",
            data=output.getvalue(),
            file_name="matchup_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Error generating table: {e}")
